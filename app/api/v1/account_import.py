from typing import List, Dict, Any
from uuid import UUID
import csv
import io

from fastapi import APIRouter, UploadFile, File, Depends, HTTPException
from sqlalchemy.orm import Session
from openpyxl import load_workbook

from app.database import get_db
from app.models.account import Account
from app.models.address import Address
from app.models.contact import Contact
from app.models.lookup_models import Department, Unit, Vertical, Location, Status


router = APIRouter(
    prefix="/accounts",
    tags=["Accounts Import"],
)


# TEMP — replace later
def get_current_user_id() -> UUID:
    return UUID("00000000-0000-0000-0000-000000000001")


# -------------------------------------------------------
# Helper: lookup by NAME
# -------------------------------------------------------
def resolve_lookup_by_name(db: Session, model, name: str | None, label: str):
    """Return model.id for given name OR None if blank."""
    if not name:
        return None

    obj = db.query(model).filter(
        model.name == name.strip(),
        model.is_deleted == False
    ).first()

    if not obj:
        raise ValueError(f"{label} '{name}' does not exist")

    return obj.id


# -------------------------------------------------------
# Helper: convert empty string → None
# -------------------------------------------------------
def null_if_empty(value):
    if value is None:
        return None
    value = str(value).strip()
    return value if value != "" else None


# -------------------------------------------------------
# Contacts parser
# -------------------------------------------------------
def parse_contacts(raw: str | None):
    """
    Format example:
    John/john@mail.com/9000; Meena/meena@mail.com/8000
    """
    if not raw:
        return []

    contacts = []
    rows = [x.strip() for x in raw.split(";") if x.strip()]

    for row in rows:
        parts = [x.strip() for x in row.split("/")]

        if len(parts) < 2:
            raise ValueError(
                f"Invalid contact format '{row}'. Expected: Name/Email/Phone"
            )

        name = parts[0]
        email = parts[1]
        phone = parts[2] if len(parts) > 2 else None

        contacts.append({"name": name, "email": email, "phone": phone})

    return contacts


# -------------------------------------------------------
# CREATE or UPDATE row
# -------------------------------------------------------
def process_import_row(db: Session, row: dict, user_id: UUID):
    """
    Handles both:
    - CREATE new account
    - UPDATE existing account (matched by Code)
    """

    # -------- Required fields --------
    name = null_if_empty(row.get("Name"))
    code = null_if_empty(row.get("Code"))

    if not name:
        raise ValueError("Name is required")
    if not code:
        raise ValueError("Code is required")

    # Does account already exist?
    account = db.query(Account).filter(
        Account.code == code,
        Account.is_deleted == False
    ).first()

    is_update = account is not None

    # -------- Optional simple fields --------
    probability = null_if_empty(row.get("Probability"))
    probability = int(probability) if probability else None

    account_partner = null_if_empty(row.get("Account Partner"))
    delivery_partner = null_if_empty(row.get("Delivery Partner"))

    # -------- Lookup fields (by NAME) --------
    department_id = resolve_lookup_by_name(db, Department, null_if_empty(row.get("Department")), "Department")
    unit_id = resolve_lookup_by_name(db, Unit, null_if_empty(row.get("Unit")), "Unit")
    vertical_id = resolve_lookup_by_name(db, Vertical, null_if_empty(row.get("Vertical")), "Vertical")
    location_id = resolve_lookup_by_name(db, Location, null_if_empty(row.get("Location")), "Location")
    status_id = resolve_lookup_by_name(db, Status, null_if_empty(row.get("Status")), "Status")

    # -------- Address fields --------
    line1 = null_if_empty(row.get("Address Line1"))
    city = null_if_empty(row.get("City"))
    country = null_if_empty(row.get("Country"))

    if not line1:
        raise ValueError("Address Line1 is required")
    if not city:
        raise ValueError("City is required")
    if not country:
        raise ValueError("Country is required")

    line2 = null_if_empty(row.get("Address Line2"))
    state = null_if_empty(row.get("State"))
    zip_code = null_if_empty(row.get("Zip"))

    # -------- Contacts --------
    contacts = parse_contacts(row.get("Contacts (Name/Email/Phone)"))

    # =====================================================
    # UPDATE EXISTING ACCOUNT
    # =====================================================
    if is_update:
        acc = account

        acc.name = name
        acc.probability = probability
        acc.account_partner = account_partner
        acc.delivery_partner = delivery_partner
        acc.department_id = department_id
        acc.unit_id = unit_id
        acc.vertical_id = vertical_id
        acc.location_id = location_id
        acc.status_id = status_id
        acc.updated_by = user_id

        # --- Update address ---
        addr = db.query(Address).filter(Address.account_id == acc.id).first()
        if addr:
            addr.addressLine1 = line1
            addr.addressLine2 = line2
            addr.city = city
            addr.state = state
            addr.countryCode = country
            addr.zip = zip_code
            addr.updated_by = user_id
        else:
            # create new if missing
            db.add(Address(
                account_id=acc.id,
                addressLine1=line1,
                addressLine2=line2,
                city=city,
                state=state,
                countryCode=country,
                zip=zip_code,
                created_by=user_id,
            ))

        # --- Replace contacts ---
        db.query(Contact).filter(Contact.account_id == acc.id).delete()

        for c in contacts:
            db.add(Contact(
                account_id=acc.id,
                name=c["name"],
                email=c["email"],
                phone=c["phone"],
                created_by=user_id,
            ))

        return "updated"

    # =====================================================
    # CREATE NEW ACCOUNT
    # =====================================================
    acc = Account(
        name=name,
        code=code,
        probability=probability,
        account_partner=account_partner,
        delivery_partner=delivery_partner,
        department_id=department_id,
        unit_id=unit_id,
        vertical_id=vertical_id,
        location_id=location_id,
        status_id=status_id,
        created_by=user_id,
    )
    db.add(acc)
    db.flush()

    # --- Create Address ---
    db.add(Address(
        account_id=acc.id,
        addressLine1=line1,
        addressLine2=line2,
        city=city,
        state=state,
        countryCode=country,
        zip=zip_code,
        created_by=user_id,
    ))

    # --- Create Contacts ---
    for c in contacts:
        db.add(Contact(
            account_id=acc.id,
            name=c["name"],
            email=c["email"],
            phone=c["phone"],
            created_by=user_id,
        ))

    return "created"


# -------------------------------------------------------
# IMPORT ENDPOINT
# -------------------------------------------------------
@router.post("/import")
async def import_accounts(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user_id: UUID = Depends(get_current_user_id),
):

    filename = file.filename.lower()
    content = await file.read()

    if not content:
        raise HTTPException(status_code=400, detail="File is empty")

    rows = []

    # CSV
    if filename.endswith(".csv"):
        decoded = content.decode("utf-8")
        reader = csv.DictReader(io.StringIO(decoded))
        for row in reader:
            rows.append({(k or "").strip(): v for k, v in row.items()})

    # XLSX
    elif filename.endswith(".xlsx"):
        wb = load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1))
        headers = [str(c.value).strip() if c.value else "" for c in header_row]

        for r in ws.iter_rows(min_row=2):
            row_dict = {}
            for i, cell in enumerate(r):
                key = headers[i] if i < len(headers) else ""
                if key:
                    row_dict[key] = cell.value
            rows.append(row_dict)
    else:
        raise HTTPException(status_code=400, detail="File must be CSV or XLSX")

    if not rows:
        raise HTTPException(status_code=400, detail="No data found in file")

    summary = {"processed": 0, "created": 0, "updated": 0, "failed": 0, "errors": []}

    for idx, row in enumerate(rows, start=2):
        try:
            result = process_import_row(db, row, current_user_id)
            db.commit()
            summary["processed"] += 1
            summary[result] += 1
        except Exception as e:
            db.rollback()
            summary["failed"] += 1
            summary["errors"].append({"row": idx, "error": str(e)})

    return summary
